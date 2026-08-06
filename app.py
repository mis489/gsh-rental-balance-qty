"""
GSH Rental Balance Qty — Web App
=================================
Upload a ZIP of rental bill PDFs -> get a formatted Excel back.

Rules Rahul asked for:
1. NEGATIVE quantities are never added into any sum. When the same party/item
   appears across multiple bills, only positive quantities are added together;
   negative lines are simply skipped (not subtracted, not summed).
2. The ZIP can contain ZIPs inside ZIPs inside RAR files, nested arbitrarily
   deep (this is how Rahul's real files are structured — outer ZIP has a ZIP
   per location, and each of those contains a RAR with the actual PDFs).
   Everything is extracted recursively before parsing.
3. Every item (SKU) has a known unit weight (from Rahul's W.T Sheet). Each
   report shows Wt (per unit) and Total Wt = Qty x Wt next to every item.

Run locally:
    pip install -r requirements.txt
    streamlit run app.py

Deploy for free (shareable link):
    Push this folder to a GitHub repo (including packages.txt), then deploy
    on share.streamlit.io.
"""
import os
import re
import glob
import shutil
import subprocess
import tempfile
import zipfile
from collections import defaultdict

import streamlit as st
from pdfminer.high_level import extract_text
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

st.set_page_config(page_title="GSH Rental Balance Qty", page_icon="📊", layout="centered")

# ════════════════════════════════════════════════════════
#  ITEM WEIGHT TABLE  (from Rahul's W.T Sheet — per-unit kg)
#  Used to compute Total Wt = Qty x Wt for every item.
#  Item names are matched in UPPERCASE, trimmed.
# ════════════════════════════════════════════════════════

WEIGHT_MAP = {
    "ACRO SPAN 2 MTR": 40.0,
    "ACRO SPAN 2.5 MTR": 40.0,
    "BASE JACK (350MM+450MM)": 3.5,
    "BASE JACK 350MM": 3.2,
    "BASE JACK 600MM": 3.5,
    "BASE JACK(650MM)": 3.5,
    "BASE-PLATE": 1.5,
    "BASEJACK-350 MM+600 MM": 3.3,
    "BASEJACK-400 MM": 3.3,
    "BASEJACK-450 MM": 3.3,
    "BASEJACK-450 MM+550 MM": 3.3,
    "CHALI": 15.0,
    "CHALI 8FT.": 20.0,
    "CHALI(6 TO 7)FT": 18.0,
    "CHALI(6 TO 8)FT": 18.0,
    "CHALI-10 FT": 27.0,
    "CHALI-6 F'T": 15.0,
    "CHANNEL -6-8 FEET": 18.0,
    "CHANNEL -7-8 FEET": 17.0,
    "CHANNEL -8-9 FEET": 20.0,
    "CHANNEL -9 FT": 20.0,
    "CHANNEL 10 FT": 24.0,
    "CHANNEL 20FT": 48.0,
    "CHANNEL 7 FT": 16.0,
    "CHANNEL 8 FT.": 18.0,
    "CHANNEL 8 TO 10 FT": 20.0,
    "CHANNEL 8-9- 10 FT": 22.0,
    "CHANNEL 9 FT.": 20.0,
    "CHANNEL 9'-10' FEET": 20.0,
    "CHANNEL-12 FT": 26.0,
    "CHANNEL-6 FT": 14.0,
    "CLAMP": 0.75,
    "CLAMP-FIXED": 0.75,
    "CLAMP-MOVING": 0.85,
    "H-FARM -2*1 MTR": 11.5,
    "H-FRAME BRASSING": 11.5,
    "INNER -3M PROP": 10.2,
    "INNER 2*3": 10.2,
    "INNER 2360 MM": 7.5,
    "INNER 2M PROP": 7.2,
    "INNER 4MTR": 24.0,
    "INNER2.5MTR": 9.0,
    "JOINTER": 0.65,
    "KAICHI-2*1 MTR": 4.0,
    "KAICHI-2.5 MTR": 5.0,
    "LEDGER - 1/2 MM": 2.0,
    "LEDGER 1.2 MTR": 4.0,
    "LEDGER 1.5 MTR": 5.0,
    "LEDGER 1.8 M": 6.0,
    "LEDGER 1M": 3.5,
    "LEDGER 1MTR": 3.5,
    "LEDGER 2M": 6.5,
    "LEDGER 2MTR": 6.5,
    "LEDGER 915MM": 3.0,
    "MS PIPE 4 MTR": 13.5,
    "MS PIPE 6 MTR": 20.0,
    "MS PIPE-2. MTR": 6.5,
    "MS PIPE-2.5 MTR": 8.5,
    "MS PIPE-3 MTR": 10.0,
    "OUTER 2": 10.2,
    "OUTER 2 MTR": 10.2,
    "OUTER 3": 15.0,
    "PROP 2*3": 20.0,
    "PROP2*2 (BASE)": 18.0,
    "PROPS 2*3 BASE": 20.0,
    "SHUTTEIRNG PLATE-3*1.8": 15.0,
    "SHUTTERIN-2*2": 18.0,
    "SHUTTERING PLATE - 3*2 (HEAVY)": 21.5,
    "SHUTTERING PLATE -4*1.5": 25.0,
    "SHUTTERING PLATE 2*1.5(WELDED)": 12.0,
    "SHUTTERING PLATE 3*14(WELDED)": 15.0,
    "SHUTTERING PLATE 3*15": 12.0,
    "SHUTTERING PLATE 3*2": 21.5,
    "SHUTTERING PLATE 3*2 (17 KGS)": 17.0,
    "SHUTTERING PLATE 3*2-1": 21.5,
    "SHUTTERING PLATE 3*2-1.5": 21.5,
    "SHUTTERING PLATE 3X0.15": 12.0,
    "SHUTTERING PLATE 3X1": 12.0,
    "SHUTTERING PLATE 3X1.5": 18.0,
    "SHUTTERING PLATE 3X9": 12.0,
    "SHUTTERING PLATE-2*1.": 12.0,
    "SHUTTERING PLATE-2*1.8": 13.0,
    "SHUTTERING PLATE-2*2": 15.0,
    "SHUTTERING PROP 2*1.5": 17.0,
    "SHUTTERING PROP 2*2": 18.0,
    "SHUTTERING PROP 2*2.5": 19.0,
    "SHUTTERING PROP 2*4": 23.0,
    "SHUTTERING PROP 3*3": 26.0,
    "SHUTTERING PROP-2*3": 21.0,
    "SHUTTERING PROP-2.5*2": 20.0,
    "SHUTTERING PROP-3*4": 28.0,
    "SIKANJA": 2.0,
    "SIKANJA -4 FT.": 2.0,
    "SIKANJA 2.5 FT.": 2.0,
    "SIKANJA 2FT": 2.0,
    "SIKANJA 3 FT.": 0.75,
    "SIKANJA 3.5 FT.": 2.0,
    "SPAN INNER": 20.0,
    "SQUARE TUBE -1.5 MTR": 5.0,
    "SQUARE TUBE -2 MTR": 6.5,
    "SQUARE TUBE -6 MTR": 20.0,
    "STANDARD 0.5 WITH 1 CUPS": 2.0,
    "STANDARD 1 MTR WITH 2 CUPS": 4.5,
    "STANDARD 1.5M WITH 3 CUPS": 7.0,
    "STANDARD 2 MTR WITH 4 CUPS": 9.0,
    "STANDARD 2.5 METER WITH 5 CUPS": 11.5,
    "STANDARD 3MTR WITH 6 CUPS": 13.2,
    "STEEL FARMA-1.2": 20.0,
    "STEEL FARMA-1.5": 22.0,
    "STEEL FORM-1.2": 22.0,
    "STEEL FORM-1.5 M": 25.0,
    "STEEL PLATE 2*1(WELDEED)": 12.0,
    "STEEL PLATE 2*18(WELDEED)": 12.0,
    "STEEL PLATE 2*2(WELDEED)": 12.0,
    "STEEL PLATE 3*.9(WELDED)": 10.0,
    "STEEL PLATE 3*1 (5.5 KGS)": 5.5,
    "STEEL PLATE 3*1(WELDED)": 12.0,
    "STEEL PLATE 3*1.15(WALDED)": 15.0,
    "STEEL PLATE 3*1.5 LIGHT WEIGHT": 15.0,
    "STEEL PLATE 3*1.5 WELDED": 15.0,
    "STEEL PLATE 3*15(WELDED)": 12.0,
    "STEEL PLATE 3*2 LIGHT WEIGHT": 17.0,
    "STEEL PLATE 3*21(WELDED)": 18.0,
    "STEEL PLATES 3* 18": 12.0,
    "STEEL PLATES 3*1.5 (14.5 KGS)": 14.5,
    "STEEL PLATES 3*18 (12 KGS)": 12.0,
    "STEEL PLATES 3*18(WELDED)": 12.0,
    "STEEL PLATES 3*2  WELDED": 21.0,
    "STEEL PLATES 3*2 (13 KGS)": 13.0,
    "STEEL PLATES 3*2 (17KG) NEW WELDED": 17.0,
    "STEEL PLATES 3*2 (17KG)WELDED": 17.0,
    "STEEL PLATES 3*2 (21 KGS)": 21.0,
    "U JACK 650 MM": 3.5,
    "UJACK (350+450)MM": 3.2,
    "UJACK 350MM": 3.2,
    "UJACK 600MM": 3.5,
    "UJACK-400 MM": 3.2,
    "UJACK-450 MM": 3.3,
    "UJACK-450+550 MM": 3.3,
    "WALKWAY CHALI-1.2 MTR": 11.5,
    "WALKWAY CHALI-1.5 MTR": 12.0,
}
def get_weight(item_name):
    """Look up per-unit weight for an item. Returns None if not in the table."""
    return WEIGHT_MAP.get(item_name.strip().upper())


# ════════════════════════════════════════════════════════
#  ARCHIVE EXTRACTION — handles ZIP-inside-ZIP-inside-RAR,
#  arbitrarily deep, exactly like Rahul's real files.
# ════════════════════════════════════════════════════════

RAR_TOOLS = [
    ["unar", "-q", "-f", "-o"],   # unar <outdir> <file>  (installed via packages.txt)
    ["unrar", "x", "-y"],          # unrar x -y <file> <outdir>/
    ["7z", "x", "-y"],             # 7z x -y -o<outdir> <file>
]

def _extract_rar(fpath, extract_dir):
    os.makedirs(extract_dir, exist_ok=True)
    # unar
    try:
        r = subprocess.run(["unar", "-q", "-f", "-o", extract_dir, fpath],
                            capture_output=True, timeout=600)
        if r.returncode == 0:
            return True
    except FileNotFoundError:
        pass
    # unrar
    try:
        r = subprocess.run(["unrar", "x", "-y", fpath, extract_dir + os.sep],
                            capture_output=True, timeout=600)
        if r.returncode == 0:
            return True
    except FileNotFoundError:
        pass
    # 7z
    try:
        r = subprocess.run(["7z", "x", "-y", f"-o{extract_dir}", fpath],
                            capture_output=True, timeout=600)
        if r.returncode == 0:
            return True
    except FileNotFoundError:
        pass
    return False

def _find_and_extract_archives(root):
    """One pass: find every .zip/.rar anywhere under root and extract it
    into a same-named '<name>_ext' folder, then delete the archive.
    Returns True if anything was extracted (caller should call again)."""
    found = False
    for dirpath, _dirnames, filenames in list(os.walk(root)):
        for fname in filenames:
            lower = fname.lower()
            if not (lower.endswith(".zip") or lower.endswith(".rar")):
                continue
            fpath = os.path.join(dirpath, fname)
            extract_dir = os.path.join(dirpath, fname.rsplit(".", 1)[0] + "_ext")
            ok = False
            if lower.endswith(".zip"):
                try:
                    with zipfile.ZipFile(fpath) as z:
                        z.extractall(extract_dir)
                    ok = True
                except Exception:
                    ok = False
            else:
                ok = _extract_rar(fpath, extract_dir)
            if ok:
                try:
                    os.remove(fpath)
                except OSError:
                    pass
                found = True
    return found

def extract_all_archives(root, max_passes=15):
    """Keep extracting nested archives until nothing is left to extract."""
    for _ in range(max_passes):
        if not _find_and_extract_archives(root):
            break


# ════════════════════════════════════════════════════════
#  PARSING
# ════════════════════════════════════════════════════════

LOCATION_MAP = {
    "GSHP": "GSHP Pune", "PUNE": "GSHP Pune",
    "GSHG": "GSHG Gurugram", "GURUGRAM": "GSHG Gurugram",
    "KSH": "KSH Karnawati", "KARNAWATI": "KSH Karnawati",
    "SSS": "SSS Ankleshwar", "ANKLESHWAR": "SSS Ankleshwar", "SUMIT": "SSS Ankleshwar",
}

def clean_location(folder_name):
    upper = folder_name.upper()
    for key, val in LOCATION_MAP.items():
        if key in upper:
            return val
    name = re.sub(r'[\-_]+', ' ', folder_name).strip()
    return name[:28]

# A real SKU name never has a standalone "-30"/"-330"-style negative number
# token (dimensions in item names are always unsigned, e.g. "2 MTR", "10 FT").
# When the PDF text has no comma between a chain of adjustment values
# (e.g. "UJACK-450 MM - 10 -30 -30 -330 -330 -1110"), that whole chunk gets
# treated as a single comma-part, and the item-name capture group ends up
# swallowing the earlier numbers. This pattern catches that so we can throw
# the entry out instead of silently recording a wrong item/qty.
_EMBEDDED_NEG_RE = re.compile(r'(?:^|\s)-\d+(?:\.\d+)?\b')

def extract_balance_qty(text):
    text = re.sub(r'\s+', ' ', text)
    match = re.search(
        r'Balance\s+Qty\s*[:\-]?\s*(.*?)(?:Cartage|Total\s+Amount|Grand\s+Total|$)',
        text, re.IGNORECASE | re.DOTALL
    )
    if not match:
        return {}, []
    balance_text = match.group(1).strip()
    items = {}
    ambiguous = []   # raw text chunks we refused to guess at
    for part in balance_text.split(','):
        part = part.strip()
        if not part:
            continue
        m = re.match(r'^(.+?)\s*-\s*([\d\.\-]+)\s*$', part)
        if m:
            item = m.group(1).strip().upper()
            # A single dash-prefixed number inside a name is normal (item
            # codes like "CHANNEL -10", "LEDGER -1.8 MTR"). Only a CHAIN of
            # 2+ such tokens means multiple un-comma'd adjustment values got
            # glued into the name — that's the actually corrupted case.
            if len(_EMBEDDED_NEG_RE.findall(item)) >= 2:
                ambiguous.append(part)
                continue
            try:
                qty = float(m.group(2))
                if item:
                    items[item] = items.get(item, 0) + qty
            except ValueError:
                pass
    return items, ambiguous

def get_party(filepath):
    base = os.path.basename(filepath).replace('.pdf', '')
    parts = base.split('_')
    if len(parts) >= 2:
        return parts[1].strip()
    return base

def get_bill_no(filepath):
    """Bill No = first underscore-separated segment of the filename,
    e.g. 'KSB00514_SUROJ BUILDCON PVT LTD...pdf' -> 'KSB00514'.
    Every bill has its own unique Bill No."""
    base = os.path.basename(filepath).replace('.pdf', '')
    parts = base.split('_')
    return parts[0].strip()


# ── Fallback: "Description For Hire Charges" table ──────────
# ONLY used when a bill has NO "Balance Qty:" line at all. When Balance Qty
# exists it is trusted as-is and this fallback is never consulted or
# cross-checked against it (per Rahul's instruction).
#
# A value line is either a bare number (the item's starting qty) or a number
# followed by a parenthetical showing how it was derived, e.g. "9(18 - 9)"
# or, when two returns land in the same period bucket, "0(9 - 7 - 2)" (2+
# deductions chained with +/-). The old version only allowed exactly one
# operator inside the parens, so a 2-deduction line like "0(9 - 7 - 2)" fell
# through as "not a number", got misread as a bogus item name, and caused
# the line before it to be wrongly flushed as the item's final value.
_HC_NUM_LINE_RE = re.compile(r'^-?[\d.,]+(\s*\(-?[\d.,]+(\s*[+\-]\s*-?[\d.,]+)+\))?$')
_HC_JUNK_NAMES = {
    "GRAND TOTAL", "TOTAL AMOUNT", "CGST", "SGST", "ROUND OFF", "CGST (9%)", "SGST (9%)",
    "CONTINUE FROM NEXT PAGE", "BANK DETAILS", "PARTY NAME", "BILLING ADDRESS", "GST NO",
    "STATE CODE", "SITE ADDRESS", "BILL NO", "FROM DATE", "DATE", "TO", "STATECODE",
    "RUNNING BILL", "DESCRIPTION FOR HIRE CHARGES",
}
_HC_BAD_NAME_RE = re.compile(r'^[A-Z]{2,6}\d{3,}$')
_HC_DATE_RE = re.compile(r'\d{1,2}-\d{1,2}-\d{2,4}')

def _hc_looks_like_item(name):
    if not name or len(name) > 45 or len(name) < 2:
        return False
    if name in _HC_JUNK_NAMES:
        return False
    if ',' in name:
        return False
    if _HC_BAD_NAME_RE.match(name.replace(' ', '')):
        return False
    if _HC_DATE_RE.search(name):
        return False
    if not re.search(r'[A-Za-z]', name):
        return False
    if 'PAGE' in name and 'OF' in name:
        return False
    return True

def extract_hire_charges_items(text):
    """Line-by-line pass over the 'Description For Hire Charges' table.
    For each item, the LAST number seen before the next item name is its
    final balance (matches how the bill itself resolves partial-period
    returns, e.g. '507(564 - 57)' -> takes 507).

    Safety net: this table's layout sometimes puts a shared Period/GST SAC
    No/Days/Number/Rate/Amount block AFTER a whole run of item name+chain
    blocks (rather than right under each item), so the last item in a run
    can accidentally swallow an unrelated number from that trailing block
    (e.g. a GST SAC code like 997313) as its "final" value. Since every
    chain is a running total that only ever decreases (starting qty, then
    successive returns/deductions), the true final value can never exceed
    the item's own starting value. Any parsed value that violates this is
    almost certainly leaked table noise, not a real quantity — it's
    discarded rather than trusted."""
    items = {}
    for m in re.finditer(r'Description For Hire Charges', text):
        start = m.end()
        header_m = re.match(r'\s*Period\s*GST SAC\s*No\s*Days\s*Number\s*Rate\s*Amount\s*', text[start:start + 200])
        if header_m:
            start += header_m.end()
        end_candidates = []
        for pat in [r'\d{1,2}\s+[A-Za-z]{3}\s*-\s*\d{1,2}\s+[A-Za-z]{3}', r'Page\s+\d+\s+of\s+\d+', r'\x0c', r'Party Name', r'Bill No']:
            mm = re.search(pat, text[start:start + 4000])
            if mm:
                end_candidates.append(start + mm.start())
        end = min(end_candidates) if end_candidates else start + 1500
        section = text[start:end]

        lines = [l.strip() for l in section.split('\n') if l.strip()]
        current_item, last_val, initial_val = None, None, None

        def flush():
            if current_item and last_val is not None and last_val != 0 and _hc_looks_like_item(current_item):
                if initial_val is not None and last_val > initial_val:
                    return  # impossible — a running deduction can't end higher than it started
                items[current_item] = items.get(current_item, 0) + last_val

        for line in lines:
            if _HC_NUM_LINE_RE.match(line):
                nm = re.match(r'^(-?[\d.]+)', line)
                if nm and current_item:
                    val = float(nm.group(1))
                    if initial_val is None:
                        initial_val = val
                    last_val = val
            else:
                flush()
                current_item, last_val, initial_val = line.upper(), None, None
        flush()
    return items

def parse_folder(root, progress_cb=None):
    bills, all_items = [], set()
    no_balance_files = []   # list of dicts: filename, path  — parsed OK but no Balance Qty line
    error_files = []        # list of dicts: filename, path, error — could not be parsed at all
    ambiguous_files = []    # list of dicts: filename, path, chunks — malformed Qty text, skipped
    recovered_count = 0     # bills where Balance Qty was missing but Hire Charges table filled it in
    pdf_files = glob.glob(os.path.join(root, '**', '*.pdf'), recursive=True)
    if not pdf_files:
        pdf_files = glob.glob(os.path.join(root, '*.pdf'))

    for idx, pdf_path in enumerate(sorted(pdf_files)):
        if progress_cb:
            progress_cb(idx + 1, len(pdf_files), os.path.basename(pdf_path))
        parent_dir = os.path.dirname(pdf_path)
        parent_name = os.path.basename(parent_dir)
        if os.path.normpath(parent_dir) == os.path.normpath(root):
            location = clean_location(os.path.basename(root))
        else:
            location = clean_location(parent_name)
        party = get_party(pdf_path)
        bill_no = get_bill_no(pdf_path)
        try:
            text = extract_text(pdf_path)
            items, ambiguous = extract_balance_qty(text)
            if ambiguous:
                ambiguous_files.append({
                    "filename": os.path.basename(pdf_path), "path": pdf_path,
                    "chunks": ambiguous,
                })
            recovered_from = None
            if not items:
                # Balance Qty line missing entirely — try the Hire Charges
                # table as a fallback. Only kicks in when Balance Qty gave
                # nothing; never overrides or cross-checks an existing value.
                fallback_items = extract_hire_charges_items(text)
                if fallback_items:
                    items = fallback_items
                    recovered_from = "hire_charges_table"
                    recovered_count += 1
                elif not ambiguous:
                    no_balance_files.append({"filename": os.path.basename(pdf_path), "path": pdf_path})
            bills.append({
                "filename": os.path.basename(pdf_path),
                "bill_no": bill_no,
                "location": location, "party": party, "items": items,
                "recovered_from": recovered_from,
            })
            all_items.update(items.keys())
        except Exception as e:
            error_files.append({"filename": os.path.basename(pdf_path), "path": pdf_path, "error": str(e)})

    return {
        "bills": bills,
        "locations": sorted(set(b['location'] for b in bills)),
        "parties": sorted(set(b['party'] for b in bills)),
        "all_items": sorted(all_items),
        "no_balance_count": len(no_balance_files),
        "no_balance_files": no_balance_files,
        "recovered_count": recovered_count,
        "ambiguous_count": len(ambiguous_files),
        "ambiguous_files": ambiguous_files,
        "errors": [f"{e['filename']}: {e['error']}" for e in error_files],
        "error_files": error_files,
        "pdf_count": len(pdf_files),
    }


def build_review_zip(no_balance_files, error_files, ambiguous_files=None):
    """Zip of PDFs where we couldn't get a Qty — for Rahul to review manually."""
    import io
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for f in no_balance_files:
            if os.path.exists(f["path"]):
                zf.write(f["path"], arcname=os.path.join("No_Balance_Qty_Found", f["filename"]))
        for f in error_files:
            if os.path.exists(f["path"]):
                zf.write(f["path"], arcname=os.path.join("Could_Not_Parse", f["filename"]))
        for f in (ambiguous_files or []):
            if os.path.exists(f["path"]):
                zf.write(f["path"], arcname=os.path.join("Ambiguous_Qty_Needs_Review", f["filename"]))
    buf.seek(0)
    return buf.getvalue()


# ════════════════════════════════════════════════════════
#  EXCEL BUILDING
#  - Negative quantities are SKIPPED when combining bills — never added.
#  - Every item shows Wt (per unit) and Total Wt (Qty x Wt).
# ════════════════════════════════════════════════════════

DARK_BLUE, MED_BLUE = "1F4E78", "2F75B5"
GREEN_DARK, GREEN_MED = "375623", "538135"
WHITE, ALT_GRAY = "FFFFFF", "F2F2F2"

def mk_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    s = Side(style='thin', color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CTR = Alignment(horizontal="center", vertical="center")
LFT = Alignment(horizontal="left", vertical="center")
WRP = Alignment(horizontal="center", vertical="center", wrap_text=True)

def title_row(ws, row, text, ncols, dark_hex, height=28):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", bold=True, size=13, color=WHITE)
    c.fill = mk_fill(dark_hex); c.alignment = CTR

def info_row(ws, row, text, ncols, med_hex, height=18):
    ws.row_dimensions[row].height = height
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", size=9, italic=True, color=WHITE)
    c.fill = mk_fill(med_hex); c.alignment = CTR

def header_cell(ws, row, col, val, dark_hex, algn=CTR):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = mk_fill(dark_hex); c.alignment = algn; c.border = thin_border()
    return c

def data_cell(ws, row, col, val, row_fill, algn=CTR, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=10); c.fill = row_fill
    c.alignment = algn; c.border = thin_border()
    if fmt:
        c.number_format = fmt
    return c

def build_dashboard(wb, sheet_title, title_text, top_parties, top_items,
                     n_bills, n_parties, n_items, grand_total, grand_weight, dark_hex, med_hex, bar_color):
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:R1")
    c = ws.cell(row=1, column=1, value=title_text)
    c.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    c.fill = mk_fill(dark_hex); c.alignment = CTR

    ws.row_dimensions[2].height = 20
    ws.merge_cells("A2:R2")
    c = ws.cell(row=2, column=1,
        value=f"Total Bills Uploaded: {n_bills}   |   Parties: {n_parties}   |   Item Types: {n_items}   |   Grand Total Qty: {int(grand_total):,}"
              f"   |   Grand Total Weight: {int(grand_weight):,} kg")
    c.font = Font(name="Arial", size=10, italic=True, color=WHITE)
    c.fill = mk_fill(med_hex); c.alignment = CTR

    ws.cell(row=3, column=1, value="Party").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=3, column=2, value="Total Qty").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=3, column=4, value="Item").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=3, column=5, value="Total Qty").font = Font(bold=True, name="Arial", size=9)

    for i, (party, qty) in enumerate(top_parties, start=4):
        ws.cell(row=i, column=1, value=party).font = Font(name="Arial", size=9)
        c = ws.cell(row=i, column=2, value=int(qty)); c.font = Font(name="Arial", size=9)
        c.number_format = '#,##0'
    for i, (item, qty) in enumerate(top_items, start=4):
        ws.cell(row=i, column=4, value=item).font = Font(name="Arial", size=9)
        c = ws.cell(row=i, column=5, value=int(qty)); c.font = Font(name="Arial", size=9)
        c.number_format = '#,##0'

    n_p = len(top_parties)
    chart1 = BarChart()
    chart1.type = "bar"; chart1.title = "Top 20 Parties — Total Qty"
    chart1.y_axis.title = "Party"; chart1.x_axis.title = "Total Qty"
    chart1.style = 10; chart1.width = 22; chart1.height = 14
    chart1.add_data(Reference(ws, min_col=2, min_row=3, max_row=3+n_p), titles_from_data=True)
    chart1.set_categories(Reference(ws, min_col=1, min_row=4, max_row=3+n_p))
    chart1.series[0].graphicalProperties.solidFill = bar_color
    ws.add_chart(chart1, "A25")

    n_i = len(top_items)
    chart2 = BarChart()
    chart2.type = "bar"; chart2.title = "Top 20 Items — Total Qty (All Parties)"
    chart2.y_axis.title = "Item"; chart2.x_axis.title = "Total Qty"
    chart2.style = 10; chart2.width = 22; chart2.height = 14
    chart2.add_data(Reference(ws, min_col=5, min_row=3, max_row=3+n_i), titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=4, min_row=4, max_row=3+n_i))
    chart2.series[0].graphicalProperties.solidFill = dark_hex
    ws.add_chart(chart2, "L25")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 14

def build_party_matrix(wb, sheet_title, title_text, party_data,
                        sorted_parties, all_items, dark_hex, med_hex,
                        party_locations=None):
    """party_locations: optional dict party -> sorted list of location names.
    When given, an extra LOCATION column is inserted right after PARTY NAME
    so it's clear which location(s) each party's bills came from."""
    ws = wb.create_sheet(title=sheet_title)
    show_loc = party_locations is not None
    name_col = 2
    loc_col = 3 if show_loc else None
    item_start_col = 4 if show_loc else 3
    qty_total_col = item_start_col + len(all_items)      # side total: sum of qty across items, per row
    weight_col = qty_total_col + 1
    total_cols = weight_col

    title_row(ws, 1, title_text, total_cols, dark_hex)
    info_row(ws, 2,
        f"Parties: {len(sorted_parties)}   |   Items: {len(all_items)}   |   Negative qty never added into totals",
        total_cols, med_hex)

    ws.row_dimensions[3].height = 50
    header_cell(ws, 3, 1, "S.No", dark_hex, CTR)
    header_cell(ws, 3, name_col, "PARTY NAME", dark_hex, Alignment(horizontal="left", vertical="center", wrap_text=True))
    if show_loc:
        header_cell(ws, 3, loc_col, "LOCATION", dark_hex, Alignment(horizontal="left", vertical="center", wrap_text=True))
    for idx, item in enumerate(all_items, start=item_start_col):
        c = ws.cell(row=3, column=idx, value=item)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = mk_fill(med_hex); c.alignment = WRP; c.border = thin_border()
    header_cell(ws, 3, qty_total_col, "TOTAL QTY", dark_hex,
                Alignment(horizontal="center", vertical="center", wrap_text=True))
    header_cell(ws, 3, weight_col, "TOTAL WEIGHT (kg)", dark_hex,
                Alignment(horizontal="center", vertical="center", wrap_text=True))

    item_col_letter_start = get_column_letter(item_start_col)
    item_col_letter_end = get_column_letter(qty_total_col - 1) if all_items else item_col_letter_start

    for row_num, party in enumerate(sorted_parties, start=1):
        er = 3 + row_num
        ws.row_dimensions[er].height = 16
        row_fill = mk_fill(ALT_GRAY) if row_num % 2 == 0 else mk_fill(WHITE)
        data_cell(ws, er, 1, row_num, row_fill, CTR)
        data_cell(ws, er, name_col, party, row_fill, LFT)
        if show_loc:
            locs = ", ".join(party_locations.get(party, []))
            data_cell(ws, er, loc_col, locs, row_fill, LFT)
        p_items = party_data.get(party, {})
        party_weight = 0.0
        for col_idx, item in enumerate(all_items, start=item_start_col):
            qty = p_items.get(item, None)
            data_cell(ws, er, col_idx, qty, row_fill, CTR, '#,##0' if qty else None)
            if qty:
                wt = get_weight(item)
                if wt is not None:
                    party_weight += qty * wt
        # Side total — sum of this party's qty across every item column
        qty_total_formula = f"=SUM({item_col_letter_start}{er}:{item_col_letter_end}{er})" if all_items else None
        data_cell(ws, er, qty_total_col, qty_total_formula, row_fill, CTR, '#,##0')
        data_cell(ws, er, weight_col, int(round(party_weight)) if party_weight else None,
                   row_fill, CTR, '#,##0')

    total_row = 3 + len(sorted_parties) + 1
    data_start, data_end = 4, 3 + len(sorted_parties)
    ws.row_dimensions[total_row].height = 18
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=item_start_col - 1)
    for col in range(1, item_start_col):
        c = ws.cell(row=total_row, column=col)
        c.fill = mk_fill(dark_hex); c.border = thin_border()
    c = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE); c.alignment = CTR

    for col_idx in range(item_start_col, weight_col + 1):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = mk_fill(dark_hex); c.alignment = CTR; c.border = thin_border()
        c.number_format = '#,##0'

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions[get_column_letter(name_col)].width = 38
    if show_loc:
        ws.column_dimensions[get_column_letter(loc_col)].width = 20
    for col_idx in range(item_start_col, item_start_col + len(all_items)):
        item = all_items[col_idx - item_start_col]
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(18, len(item) * 0.85))
    ws.column_dimensions[get_column_letter(qty_total_col)].width = 14
    ws.column_dimensions[get_column_letter(weight_col)].width = 16
    ws.freeze_panes = f"{get_column_letter(item_start_col)}4"

def build_location_tab(wb, sheet_title, location_name, items_dict, dark_hex, med_hex, n_bills=None):
    ws = wb.create_sheet(title=sheet_title[:31])
    title_row(ws, 1, location_name, 5, dark_hex)
    bills_txt = f"Bills: {n_bills}   |   " if n_bills is not None else ""
    info_row(ws, 2, f"{bills_txt}Items shown: {len(items_dict)}   |   Negative qty never added into totals", 5, med_hex)

    ws.row_dimensions[3].height = 20
    header_cell(ws, 3, 1, "S.No", dark_hex, CTR)
    header_cell(ws, 3, 2, "Item Name", dark_hex, Alignment(horizontal="left", vertical="center"))
    header_cell(ws, 3, 3, "Qty", dark_hex, CTR)
    header_cell(ws, 3, 4, "Wt (per unit)", dark_hex, CTR)
    header_cell(ws, 3, 5, "Total Wt", dark_hex, CTR)

    sorted_items = sorted(items_dict.items(), key=lambda x: x[0])  # A to Z by item name
    for row_num, (item, qty) in enumerate(sorted_items, start=1):
        er = 3 + row_num
        ws.row_dimensions[er].height = 16
        row_fill = mk_fill(ALT_GRAY) if row_num % 2 == 0 else mk_fill(WHITE)
        wt = get_weight(item)
        data_cell(ws, er, 1, row_num, row_fill, CTR)
        data_cell(ws, er, 2, item, row_fill, LFT)
        data_cell(ws, er, 3, qty, row_fill, CTR, '#,##0')
        data_cell(ws, er, 4, wt, row_fill, CTR, '#,##0.00' if wt else None)
        data_cell(ws, er, 5, round(qty * wt, 2) if wt is not None else None, row_fill, CTR, '#,##0')

    n = len(sorted_items)
    total_r = 3 + n + 1
    ws.row_dimensions[total_r].height = 18
    for col, val in [(1, "TOTAL"), (2, ""), (4, "")]:
        c = ws.cell(row=total_r, column=col, value=val)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = mk_fill(dark_hex); c.border = thin_border(); c.alignment = CTR
    c = ws.cell(row=total_r, column=3, value=f"=SUM(C4:C{3+n})")
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = mk_fill(dark_hex); c.border = thin_border(); c.alignment = CTR
    c.number_format = '#,##0'
    cw = ws.cell(row=total_r, column=5, value=f"=SUM(E4:E{3+n})")
    cw.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    cw.fill = mk_fill(dark_hex); cw.border = thin_border(); cw.alignment = CTR
    cw.number_format = '#,##0'

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A4"


def build_missing_weight_tab(wb, items_list, dark_hex, med_hex):
    ws = wb.create_sheet(title="Missing Weight")
    title_row(ws, 1, "SKUs With No Weight Match", 2, dark_hex)
    info_row(ws, 2, f"{len(items_list)} items — add these to the weight list so Total Wt can be calculated", 2, med_hex)

    ws.row_dimensions[3].height = 20
    header_cell(ws, 3, 1, "S.No", dark_hex, CTR)
    header_cell(ws, 3, 2, "Item Name", dark_hex, Alignment(horizontal="left", vertical="center"))

    for row_num, item in enumerate(items_list, start=1):
        er = 3 + row_num
        ws.row_dimensions[er].height = 16
        row_fill = mk_fill(ALT_GRAY) if row_num % 2 == 0 else mk_fill(WHITE)
        data_cell(ws, er, 1, row_num, row_fill, CTR)
        data_cell(ws, er, 2, item, row_fill, LFT)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 46
    ws.freeze_panes = "A4"


def build_bill_wise_tab(wb, sheet_title, title_text, bills, all_items, dark_hex, med_hex):
    """One row PER BILL (not aggregated) — S.No, Bill No, Party Name, Location,
    then that bill's own item quantities. Negative/zero qty is blanked out on
    a per-row basis, same convention as everywhere else in the report."""
    ws = wb.create_sheet(title=sheet_title[:31])
    bill_col = 2
    party_col = 3
    loc_col = 4
    item_start_col = 5
    qty_total_col = item_start_col + len(all_items)
    weight_col = qty_total_col + 1
    total_cols = weight_col

    title_row(ws, 1, title_text, total_cols, dark_hex)
    info_row(ws, 2,
        f"Bills: {len(bills)}   |   Items: {len(all_items)}   |   Negative qty never added into totals",
        total_cols, med_hex)

    ws.row_dimensions[3].height = 50
    header_cell(ws, 3, 1, "S.No", dark_hex, CTR)
    header_cell(ws, 3, bill_col, "BILL NO", dark_hex, CTR)
    header_cell(ws, 3, party_col, "PARTY NAME", dark_hex, Alignment(horizontal="left", vertical="center", wrap_text=True))
    header_cell(ws, 3, loc_col, "LOCATION", dark_hex, Alignment(horizontal="left", vertical="center", wrap_text=True))
    for idx, item in enumerate(all_items, start=item_start_col):
        c = ws.cell(row=3, column=idx, value=item)
        c.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        c.fill = mk_fill(med_hex); c.alignment = WRP; c.border = thin_border()
    header_cell(ws, 3, qty_total_col, "TOTAL QTY", dark_hex,
                Alignment(horizontal="center", vertical="center", wrap_text=True))
    header_cell(ws, 3, weight_col, "TOTAL WEIGHT (kg)", dark_hex,
                Alignment(horizontal="center", vertical="center", wrap_text=True))

    item_col_letter_start = get_column_letter(item_start_col)
    item_col_letter_end = get_column_letter(qty_total_col - 1) if all_items else item_col_letter_start

    sorted_bills = sorted(bills, key=lambda b: (b['location'], b['bill_no']))

    for row_num, bill in enumerate(sorted_bills, start=1):
        er = 3 + row_num
        ws.row_dimensions[er].height = 16
        row_fill = mk_fill(ALT_GRAY) if row_num % 2 == 0 else mk_fill(WHITE)
        data_cell(ws, er, 1, row_num, row_fill, CTR)
        data_cell(ws, er, bill_col, bill['bill_no'], row_fill, CTR)
        data_cell(ws, er, party_col, bill['party'], row_fill, LFT)
        data_cell(ws, er, loc_col, bill['location'], row_fill, LFT)
        bill_weight = 0.0
        for col_idx, item in enumerate(all_items, start=item_start_col):
            qty = bill['items'].get(item)
            if qty is not None and qty <= 0:
                qty = None   # negative/zero never shown — same rule as rest of the report
            data_cell(ws, er, col_idx, qty, row_fill, CTR, '#,##0' if qty else None)
            if qty:
                wt = get_weight(item)
                if wt is not None:
                    bill_weight += qty * wt
        qty_total_formula = f"=SUM({item_col_letter_start}{er}:{item_col_letter_end}{er})" if all_items else None
        data_cell(ws, er, qty_total_col, qty_total_formula, row_fill, CTR, '#,##0')
        data_cell(ws, er, weight_col, int(round(bill_weight)) if bill_weight else None,
                   row_fill, CTR, '#,##0')

    total_row = 3 + len(sorted_bills) + 1
    data_start, data_end = 4, 3 + len(sorted_bills)
    ws.row_dimensions[total_row].height = 18
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=item_start_col - 1)
    for col in range(1, item_start_col):
        c = ws.cell(row=total_row, column=col)
        c.fill = mk_fill(dark_hex); c.border = thin_border()
    c = ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE); c.alignment = CTR

    for col_idx in range(item_start_col, weight_col + 1):
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})")
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = mk_fill(dark_hex); c.alignment = CTR; c.border = thin_border()
        c.number_format = '#,##0'

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions[get_column_letter(bill_col)].width = 16
    ws.column_dimensions[get_column_letter(party_col)].width = 34
    ws.column_dimensions[get_column_letter(loc_col)].width = 18
    for col_idx in range(item_start_col, item_start_col + len(all_items)):
        item = all_items[col_idx - item_start_col]
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(18, len(item) * 0.85))
    ws.column_dimensions[get_column_letter(qty_total_col)].width = 14
    ws.column_dimensions[get_column_letter(weight_col)].width = 16
    ws.freeze_panes = f"{get_column_letter(item_start_col)}4"


def build_excel(data, out_path, title):
    bills = data['bills']
    locations = data['locations']
    parties = data['parties']
    all_items = data['all_items']

    # ── Aggregate by party — NEGATIVE QTY SKIPPED, NEVER ADDED ──
    party_data = defaultdict(lambda: defaultdict(float))
    for bill in bills:
        for item, qty in bill['items'].items():
            if qty > 0:
                party_data[bill['party']][item] += qty

    # ── Aggregate by location — same rule ──
    loc_data = defaultdict(lambda: defaultdict(float))
    for bill in bills:
        for item, qty in bill['items'].items():
            if qty > 0:
                loc_data[bill['location']][item] += qty

    # ── Which location(s) each party has bills from ──
    party_locations = defaultdict(set)
    for bill in bills:
        party_locations[bill['party']].add(bill['location'])
    party_locations = {p: sorted(locs) for p, locs in party_locations.items()}

    # ── Aggregate by (location, party) — for the per-location Party Wise tabs ──
    loc_party_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for bill in bills:
        for item, qty in bill['items'].items():
            if qty > 0:
                loc_party_data[bill['location']][bill['party']][item] += qty

    # ── Real item universe = only names that actually carry a positive qty
    # somewhere in the report. Garbage/junk strings that only ever showed up
    # with a negative (skipped) qty — e.g. messy PDF text like a run of
    # "-30 -30 -330" adjustments glued onto an item name — never make it in,
    # so they can't pollute the Party Wise columns, item counts, or the
    # Missing Weight tab. ──
    real_items = set()
    for items in party_data.values():
        real_items.update(items.keys())
    all_items = sorted(real_items)

    def top_parties():
        totals = {p: sum(items.values()) for p, items in party_data.items() if sum(items.values()) > 0}
        top20 = sorted(totals.items(), key=lambda x: x[1], reverse=True)[:20]
        return sorted(top20, key=lambda x: x[0])  # biggest 20 selected, then shown A to Z

    def top_items():
        totals = defaultdict(float)
        for p, items in party_data.items():
            for item, qty in items.items():
                totals[item] += qty
        top20 = sorted(totals.items(), key=lambda x: x[1], reverse=True)[:20]
        return sorted(top20, key=lambda x: x[0])  # biggest 20 selected, then shown A to Z

    grand_total = sum(qty for p, items in party_data.items() for qty in items.values())
    grand_weight = sum(
        qty * get_weight(item) for p, items in party_data.items()
        for item, qty in items.items() if get_weight(item) is not None
    )

    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(
        wb, "Dashboard", f"📊 {title}",
        top_parties(), top_items(),
        len(bills), len(parties), len(all_items), grand_total, grand_weight,
        DARK_BLUE, MED_BLUE, "2F75B5"
    )

    build_party_matrix(
        wb, "Party Wise", f"{title}  —  PARTY WISE (All Locations)",
        {p: dict(v) for p, v in party_data.items()},
        sorted(parties), all_items, DARK_BLUE, MED_BLUE,
        party_locations=party_locations
    )

    # ── NEW: Bill Wise Stock Report — one row per individual bill, with its
    # own unique Bill No, Party Name, and that bill's own (not aggregated)
    # item quantities. Added per Rahul's request — everything above/below
    # this stays exactly as it was. ──
    build_bill_wise_tab(
        wb, "Bill Wise", f"{title}  —  BILL WISE STOCK REPORT",
        bills, all_items, DARK_BLUE, MED_BLUE
    )

    bills_per_location = defaultdict(int)
    for bill in bills:
        bills_per_location[bill['location']] += 1

    for loc in sorted(locations):
        items_dict = dict(loc_data[loc])
        build_location_tab(wb, loc, loc, items_dict, DARK_BLUE, MED_BLUE, n_bills=bills_per_location[loc])

    # ── Per-location Party Wise tabs — same matrix, filtered to one location ──
    for loc in sorted(locations):
        loc_parties = sorted(loc_party_data[loc].keys())
        loc_items = sorted({item for p in loc_parties for item in loc_party_data[loc][p].keys()})
        build_party_matrix(
            wb, f"PW - {loc}"[:31], f"{title}  —  PARTY WISE  ({loc})",
            {p: dict(v) for p, v in loc_party_data[loc].items()},
            loc_parties, loc_items, DARK_BLUE, MED_BLUE
        )

    unmatched = sorted({item for item in all_items if get_weight(item) is None})
    if unmatched:
        build_missing_weight_tab(wb, unmatched, DARK_BLUE, MED_BLUE)

    wb.save(out_path)
    return unmatched


# ════════════════════════════════════════════════════════
#  STREAMLIT UI
# ════════════════════════════════════════════════════════

st.title("📊 GSH Rental Balance Qty")
st.caption("ZIP upload karo (ZIP ke andar ZIP/RAR bhi chalega, kitni bhi layers deep). "
           "Excel me har item ka Qty, Weight aur Total Weight milega. "
           "Negative qty kabhi sum me add nahi hoti.")

uploaded_zip = st.file_uploader("Rental bills ka ZIP file upload karo", type=["zip"])
report_title = st.text_input("Report title", value="GSH Rental Balance Qty")

if uploaded_zip is not None:
    if st.button("🚀 Excel Banao", type="primary"):
        with tempfile.TemporaryDirectory() as tmpdir:
            zip_path = os.path.join(tmpdir, "input.zip")
            with open(zip_path, "wb") as f:
                f.write(uploaded_zip.getbuffer())

            bills_dir = os.path.join(tmpdir, "bills")
            with zipfile.ZipFile(zip_path, "r") as z:
                z.extractall(bills_dir)

            with st.spinner("Andar ke ZIP/RAR files khol raha hoon (jitni bhi layers hon)..."):
                extract_all_archives(bills_dir)

            progress_bar = st.progress(0)
            status_text = st.empty()

            def progress_cb(done, total, name):
                progress_bar.progress(done / total)
                status_text.text(f"Parsing {done}/{total}: {name}")

            with st.spinner("PDFs parse ho rahe hain..."):
                data = parse_folder(bills_dir, progress_cb)

            status_text.empty()
            progress_bar.empty()

            if data['pdf_count'] == 0:
                st.session_state.pop('result', None)
                st.error("Koi PDF nahi mila. ZIP ke andar ka structure check karo.")
            else:
                out_path = os.path.join(tmpdir, "Rental_Balance_Qty.xlsx")
                unmatched = build_excel(data, out_path, report_title)
                with open(out_path, "rb") as f:
                    excel_bytes = f.read()

                review_files = (
                    data.get('no_balance_files', [])
                    + data.get('error_files', [])
                    + data.get('ambiguous_files', [])
                )
                review_zip_bytes = None
                if review_files:
                    review_zip_bytes = build_review_zip(
                        data.get('no_balance_files', []), data.get('error_files', []),
                        data.get('ambiguous_files', [])
                    )

                # Save everything needed to render the results — this survives
                # the rerun that happens when a download button is clicked.
                st.session_state['result'] = {
                    "data_summary": {
                        "pdf_count": data['pdf_count'],
                        "locations": data['locations'],
                        "parties": data['parties'],
                        "all_items": data['all_items'],
                        "recovered_count": data.get('recovered_count', 0),
                        "no_balance_count": data['no_balance_count'],
                        "ambiguous_count": data.get('ambiguous_count', 0),
                        "errors": data['errors'],
                    },
                    "unmatched": unmatched,
                    "excel_bytes": excel_bytes,
                    "review_zip_bytes": review_zip_bytes,
                    "review_count": len(review_files),
                }

# ── Render results (persists across download-button reruns) ──
if 'result' in st.session_state:
    r = st.session_state['result']
    s = r['data_summary']
    st.success(
        f"✅ Done — PDFs: {s['pdf_count']} | Locations: {len(s['locations'])} "
        f"| Parties: {len(s['parties'])} | Items: {len(s['all_items'])}"
    )
    if s['locations']:
        st.write("Locations: " + ", ".join(s['locations']))
    if s['no_balance_count']:
        st.info(f"{s['no_balance_count']} bills me kahi se bhi Qty nahi mili (final bills — skip kiye).")
    if s.get('ambiguous_count'):
        st.warning(f"⚠️ {s['ambiguous_count']} bills me Qty text confusing tha (bina comma ke jude hue numbers) — safety ke liye skip kar diya, download me 'Ambiguous_Qty_Needs_Review' folder me milega, manually check karo.")
    if s['errors']:
        with st.expander(f"⚠️ {len(s['errors'])} PDFs parse nahi hue — dekho"):
            for e in s['errors'][:50]:
                st.text(e)

    if r['unmatched']:
        with st.expander(f"⚖️ {len(r['unmatched'])} items ki weight table me nahi mili (Total Wt blank rahega)"):
            for item in r['unmatched']:
                st.text(item)

    st.download_button(
        "⬇️ Excel Download Karo",
        data=r['excel_bytes'],
        file_name="Rental_Balance_Qty.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        key="dl_excel",
    )

    if r['review_zip_bytes']:
        st.download_button(
            f"⬇️ Un {r['review_count']} Bills ki PDFs Download Karo (jinki Qty nahi mili)",
            data=r['review_zip_bytes'],
            file_name="Bills_Needing_Review.zip",
            mime="application/zip",
            key="dl_review_zip",
        )

st.divider()
st.caption("Data sirf is session me process hota hai, kahi save nahi hota. Har upload ek naya, alag run hai.")
